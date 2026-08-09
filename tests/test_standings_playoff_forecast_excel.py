"""Workbook contract tests for the standings/playoff forecast Excel layer."""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from tests.test_standings_playoff_forecast_outputs import (  # noqa: E402
    _inputs,
    _write_fixture_bundle,
)


EXPECTED_SHEETS = [
    "Dashboard",
    "Current Standings",
    "Remaining Schedule",
    "Forecast",
    "Broadcast Insights",
    "Team Games Source",
    "Model Notes",
]


def _team_games() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "season": 2031,
                "season_type": "Regular Season",
                "game_id": "g1",
                "game_date": pd.Timestamp("2031-06-01"),
                "team_id": "B",
                "team_abbreviation": "BRV",
                "opponent_id": "A",
                "opponent_abbreviation": "ALP",
                "home_away": "away",
                "win": 0,
                "loss": 1,
                "points_for": 70,
                "points_against": 80,
                "margin": -10,
                "wins_to_date": 0,
                "losses_to_date": 1,
                "win_pct_to_date": 0.0,
                "point_diff_to_date": -10,
                "source_game_completed": True,
                "source_team_box_path": "literal/team_box.parquet",
            },
            {
                "season": 2031,
                "season_type": "Regular Season",
                "game_id": "g1",
                "game_date": pd.Timestamp("2031-06-01"),
                "team_id": "A",
                "team_abbreviation": "ALP",
                "opponent_id": "B",
                "opponent_abbreviation": "BRV",
                "home_away": "home",
                "win": 1,
                "loss": 0,
                "points_for": 80,
                "points_against": 70,
                "margin": 10,
                "wins_to_date": 1,
                "losses_to_date": 0,
                "win_pct_to_date": 1.0,
                "point_diff_to_date": 10,
                "source_game_completed": True,
                "source_team_box_path": "literal/team_box.parquet",
            },
        ]
    )


def _literal_bundle(root: Path, *, history_available: bool = True):
    cfg, model_cfg, bundle = _inputs(root)
    # Deliberately non-round probabilities prove the workbook is a view of the
    # validated bundle rather than a second simulation/calculation surface.
    bundle.simulation_result.forecast_summary.loc[:, "playoff_probability"] = [
        0.731,
        0.269,
    ]
    bundle.simulation_result.forecast_summary.loc[:, "top4_probability"] = [
        0.731,
        0.269,
    ]
    bundle.simulation_result.forecast_summary.loc[:, "home_court_probability"] = [
        0.731,
        0.269,
    ]
    bundle.simulation_result.forecast_summary.loc[:, "out_probability"] = [
        0.269,
        0.731,
    ]
    bundle.simulation_result.forecast_summary.loc[:, "expected_final_rank"] = [
        1.269,
        1.731,
    ]
    bundle.simulation_result.forecast_summary.loc[:, "rank_1_probability"] = [
        0.731,
        0.269,
    ]
    bundle.simulation_result.forecast_summary.loc[:, "rank_2_probability"] = [
        0.269,
        0.731,
    ]
    bundle.simulation_result.rank_probability_matrix.loc[:, "probability"] = [
        0.731,
        0.269,
        0.269,
        0.731,
    ]
    categories = [
        "leader_race",
        "playoff_cutline",
        "top4_race",
        "tiebreak_watch",
        "remaining_sos",
        "most_likely_riser",
        "most_vulnerable_seed",
        "recent_form",
        "high_leverage_game",
        "historical_context",
    ]
    base_insight = bundle.broadcast_insights.iloc[0].to_dict()
    insights = []
    for priority, category in enumerate(categories, start=1):
        row = dict(base_insight)
        row.update(
            priority=priority,
            category=category,
            quick_read_snippet=f"Literal story {priority}.",
        )
        insights.append(row)
    bundle = replace(bundle, broadcast_insights=pd.DataFrame(insights))
    if not history_available:
        bundle.historical_context.drop(bundle.historical_context.index, inplace=True)
    processed_root = _write_fixture_bundle(root, cfg, model_cfg, bundle)
    return cfg, processed_root


def _header_map(sheet) -> dict[str, int]:
    return {cell.value: cell.column for cell in sheet[2] if cell.value is not None}


class ExcelRendererTests(unittest.TestCase):
    def test_renders_exact_workbook_contract_without_recalculating_forecast(self):
        from standings_playoff_forecast.render_excel import render_excel

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            cfg, processed_root = _literal_bundle(root)
            workbook_path = render_excel(processed_root, _team_games(), cfg=cfg)

            expected_path = (
                Path(cfg.output_root)
                / "deliverables"
                / "season=2031"
                / "latest"
                / "wnba_standings_playoff_forecast.xlsx"
            )
            self.assertEqual(workbook_path, expected_path)
            self.assertTrue(workbook_path.is_file())

            workbook = load_workbook(workbook_path, data_only=False)
            self.assertEqual(workbook.sheetnames, EXPECTED_SHEETS)
            self.assertTrue(
                workbook["Dashboard"].sheet_properties.pageSetUpPr.fitToPage
            )
            self.assertEqual(workbook["Dashboard"].page_setup.fitToWidth, 1)
            self.assertEqual(workbook["Model Notes"].page_setup.fitToHeight, 1)
            dashboard_text = {
                cell.value
                for row in workbook["Dashboard"].iter_rows()
                for cell in row
                if isinstance(cell.value, str)
            }
            for label in (
                "Current Leader",
                "2031 Playoff Cutline (1st / 2nd)",
                "Projected Playoff Field",
                "Top-4 / Home-Court Race",
                "Most Uncertain Seed",
                "Remaining SOS Extremes",
                "Highest-Leverage Games",
                "Headline Storylines",
                "Historical Cutline Comparison",
            ):
                self.assertIn(label, dashboard_text)

            forecast = workbook["Forecast"]
            headers = _header_map(forecast)
            self.assertIn("Rank 1 %", headers)
            self.assertIn("Rank 2 %", headers)
            self.assertIn("OUT %", headers)
            self.assertEqual(forecast.cell(3, headers["Team"]).value, "ALP")
            supplied_probability = pd.read_csv(
                processed_root / "forecast_summary.csv"
            ).sort_values(["current_rank", "team_id"], kind="stable").iloc[0][
                "playoff_probability"
            ]
            self.assertEqual(
                forecast.cell(3, headers["Playoff %"]).value,
                supplied_probability,
            )
            self.assertEqual(forecast.cell(3, headers["Playoff %"]).number_format, "0.0%")

            remaining = workbook["Remaining Schedule"]
            remaining_headers = _header_map(remaining)
            self.assertIsInstance(
                remaining.cell(3, remaining_headers["Date"]).value,
                pd.Timestamp.to_pydatetime(pd.Timestamp("2031-06-03")).__class__,
            )
            self.assertEqual(remaining.freeze_panes, "A3")
            self.assertEqual(remaining.auto_filter.ref, f"A2:S{remaining.max_row}")
            self.assertTrue(remaining.tables)

            source = workbook["Team Games Source"]
            source_headers = _header_map(source)
            self.assertEqual(source.cell(3, source_headers["Team ID"]).value, "A")
            self.assertEqual(source.freeze_panes, "F3")
            self.assertTrue(source.tables)

            formulas = [
                cell.value
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ]
            self.assertTrue(formulas)
            self.assertTrue(any("='Current Standings'!" in value for value in formulas))
            for formula in formulas:
                self.assertNotIn("#REF!", formula)

            data_only = load_workbook(workbook_path, data_only=True)
            self.assertEqual(data_only.sheetnames, EXPECTED_SHEETS)
            self.assertEqual(
                data_only["Forecast"].cell(3, headers["Playoff %"]).value,
                supplied_probability,
            )
            error_tokens = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
            for opened in (workbook, data_only):
                errors = [
                    (sheet.title, cell.coordinate, cell.value)
                    for sheet in opened.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str)
                    and any(token in cell.value for token in error_tokens)
                ]
                self.assertEqual(errors, [])
            self.assertEqual(workbook["Broadcast Insights"].max_row, 12)

    def test_history_unavailable_is_explicit_and_repeated_write_is_safe(self):
        from standings_playoff_forecast.render_excel import render_excel

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            cfg, processed_root = _literal_bundle(root, history_available=False)
            first = render_excel(processed_root, _team_games(), cfg=cfg)
            second = render_excel(processed_root, _team_games(), cfg=cfg)
            self.assertEqual(first, second)
            workbook = load_workbook(second, data_only=False)
            values = [
                cell.value
                for row in workbook["Dashboard"].iter_rows()
                for cell in row
            ]
            self.assertIn("Historical context unavailable for this run.", values)
            self.assertFalse(list(second.parent.glob(f".{second.name}.*.tmp")))

    def test_schema_failure_does_not_publish_partial_workbook(self):
        from standings_playoff_forecast.render_excel import render_excel

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            cfg, processed_root = _literal_bundle(root)
            forecast_path = processed_root / "forecast_summary.csv"
            forecast = pd.read_csv(forecast_path).drop(columns=["playoff_probability"])
            forecast.to_csv(forecast_path, index=False)

            final_path = (
                Path(cfg.output_root)
                / "deliverables"
                / "season=2031"
                / "latest"
                / "wnba_standings_playoff_forecast.xlsx"
            )
            with self.assertRaisesRegex(ValueError, "forecast_summary.*playoff_probability"):
                render_excel(processed_root, _team_games(), cfg=cfg)
            self.assertFalse(final_path.exists())
            self.assertFalse(final_path.parent.exists())

    def test_manifest_and_model_notes_preserve_provenance(self):
        from standings_playoff_forecast.render_excel import render_excel

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            cfg, processed_root = _literal_bundle(root)
            manifest = json.loads((processed_root / "run_manifest.json").read_text())
            workbook_path = render_excel(processed_root, _team_games(), cfg=cfg)
            workbook = load_workbook(workbook_path, data_only=False)
            notes = workbook["Model Notes"]
            note_rows = {
                notes.cell(row, 1).value: notes.cell(row, 2).value
                for row in range(3, notes.max_row + 1)
            }
            self.assertEqual(note_rows["Cutoff date"], manifest["cutoff_date"])
            self.assertEqual(note_rows["Model version"], manifest["model_version"])
            self.assertEqual(note_rows["Simulation count"], manifest["simulation_count"])
            self.assertEqual(note_rows["Random seed"], manifest["random_seed"])
            self.assertIn("machine-readable forecast bundle", note_rows["Probability source"])


if __name__ == "__main__":
    unittest.main()
