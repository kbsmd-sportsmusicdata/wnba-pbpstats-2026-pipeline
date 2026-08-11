"""Render the validated standings forecast bundle as an Excel companion analysis."""

from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path
from typing import Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .outputs import CSV_FILENAMES, PAYLOAD_KEYS
from .render_markdown import CURRENT_500_METHOD, CURRENT_STANDINGS_METHOD
from .team_game_layer import normalize_id


WORKBOOK_FILENAME = "wnba_standings_playoff_forecast.xlsx"
SHEET_ORDER = (
    "Dashboard",
    "Current Standings",
    "Remaining Schedule",
    "Forecast",
    "Broadcast Insights",
    "Team Games Source",
    "Model Notes",
)

NAVY = "13233F"
BLUE = "1E5A93"
SKY = "DCEAF7"
PALE_BLUE = "EDF4FA"
GOLD = "E6B44A"
PALE_GOLD = "FAF1D9"
GREEN = "2F7D5A"
PALE_GREEN = "E5F2EC"
RED = "B44646"
PALE_RED = "F7E5E5"
GRAY = "5D6877"
LIGHT_GRAY = "E7EBF0"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
SECTION_FONT = Font(name="Aptos Display", size=11, bold=True, color=WHITE)
HEADER_FONT = Font(name="Aptos", size=10, bold=True, color=WHITE)
BODY_FONT = Font(name="Aptos", size=10, color="1F2933")
SUBTLE_FONT = Font(name="Aptos", size=9, color=GRAY)
THIN_GRAY = Side(style="thin", color="CBD2DA")


REQUIRED_COLUMNS = {
    "current_standings": {
        "team_id",
        "team_abbreviation",
        "current_rank",
        "games_played",
        "wins",
        "losses",
        "win_pct",
        "point_differential",
        "games_back",
        "home_record",
        "road_record",
        "last10_record",
        "current_streak_label",
        "conference_record",
        "record_vs_current_500_plus",
    },
    "head_to_head": {
        "team_id",
        "opponent_id",
        "games_played",
        "wins",
        "losses",
        "win_pct",
        "point_differential",
    },
    "team_strength": {
        "team_id",
        "season_net_rating",
        "recent_net_rating",
        "composite_strength",
        "pbpstats_snapshot_available",
        "pbpstats_snapshot_as_of",
        "pbpstats_snapshot_safe_for_cutoff",
    },
    "remaining_schedule": {
        "game_id",
        "game_date",
        "home_id",
        "away_id",
        "home_rest_days",
        "away_rest_days",
        "home_b2b",
        "away_b2b",
    },
    "matchup_probabilities": {
        "game_id",
        "game_date",
        "home_id",
        "away_id",
        "home_rest_days",
        "away_rest_days",
        "home_b2b",
        "away_b2b",
        "home_win_probability",
        "away_win_probability",
        "expected_home_margin",
    },
    "forecast_summary": {
        "season",
        "cutoff_date",
        "team_id",
        "team_abbreviation",
        "current_rank",
        "projected_wins_mean",
        "wins_p10",
        "wins_p50",
        "wins_p90",
        "expected_final_rank",
        "playoff_probability",
        "top4_probability",
        "home_court_probability",
        "out_probability",
        "remaining_sos_rank",
    },
    "rank_probability_matrix": {
        "season",
        "team_id",
        "team_abbreviation",
        "final_rank",
        "probability",
    },
    "playoff_leverage_games": {
        "game_id",
        "game_date",
        "home_id",
        "away_id",
        "leverage_score",
        "leverage_label",
        "direct_h2h_tiebreak_flag",
    },
    "historical_context": {
        "context_level",
        "metric",
        "season",
        "value",
        "availability_status",
    },
    "broadcast_insights": {
        "priority",
        "category",
        "team_or_race",
        "data_point",
        "quick_read_snippet",
        "pregame_trigger",
        "halftime_checkpoint",
        "fourth_quarter_checkpoint",
        "why_it_matters",
        "source_fields",
    },
}

TEAM_GAME_REQUIRED = {
    "season",
    "game_id",
    "game_date",
    "team_id",
    "opponent_id",
    "win",
    "loss",
    "points_for",
    "points_against",
    "margin",
    "source_game_completed",
    "source_team_box_path",
}


def _require_columns(frame: pd.DataFrame, required: set[str], name: str) -> None:
    missing = sorted(required.difference(frame.columns))
    if missing:
        raise ValueError(f"{name} is missing required columns: {', '.join(missing)}")


def _load_team_games(team_games: pd.DataFrame | str | Path) -> pd.DataFrame:
    if isinstance(team_games, pd.DataFrame):
        frame = team_games.copy()
    else:
        path = Path(team_games)
        if not path.is_file():
            raise FileNotFoundError(f"team-game source does not exist: {path}")
        if path.suffix.lower() != ".parquet":
            raise ValueError("team-game source path must be a parquet file")
        frame = pd.read_parquet(path)
    _require_columns(frame, TEAM_GAME_REQUIRED, "team_games")
    if frame.empty:
        raise ValueError("team_games must contain normalized directional rows")
    return frame


def _normalize_ids(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in result.columns:
        if column == "game_id" or column.endswith("_id"):
            result[column] = result[column].map(normalize_id)
    return result


def _load_inputs(
    processed_root: Path, team_games: pd.DataFrame | str | Path, cfg: object
) -> tuple[dict[str, pd.DataFrame], pd.DataFrame, dict, dict]:
    if not processed_root.is_dir():
        raise FileNotFoundError(f"processed forecast root does not exist: {processed_root}")
    missing_files = [
        filename
        for filename in (*CSV_FILENAMES.values(), "forecast_payload.json", "run_manifest.json")
        if not (processed_root / filename).is_file()
    ]
    if missing_files:
        raise FileNotFoundError(
            "processed forecast bundle is missing: " + ", ".join(sorted(missing_files))
        )

    frames = {
        name: _normalize_ids(pd.read_csv(processed_root / filename))
        for name, filename in CSV_FILENAMES.items()
    }
    for name, required in REQUIRED_COLUMNS.items():
        _require_columns(frames[name], required, name)
    for rank in range(1, int(cfg.team_count) + 1):
        _require_columns(
            frames["forecast_summary"],
            {f"rank_{rank}_probability"},
            "forecast_summary",
        )

    manifest = json.loads((processed_root / "run_manifest.json").read_text(encoding="utf-8"))
    payload = json.loads((processed_root / "forecast_payload.json").read_text(encoding="utf-8"))
    if set(payload) != PAYLOAD_KEYS:
        raise ValueError("forecast_payload has unknown or missing top-level keys")
    if payload.get("metadata") != manifest:
        raise ValueError("forecast_payload metadata must match run_manifest")
    if int(manifest.get("season", -1)) != int(cfg.season):
        raise ValueError("run_manifest season does not match configured season")

    current = frames["current_standings"]
    forecast = frames["forecast_summary"]
    if set(current["team_id"]) != set(forecast["team_id"]):
        raise ValueError("forecast_summary team universe must match current_standings")
    if len(current) != int(cfg.team_count):
        raise ValueError("current_standings team count does not match configured season")
    ranks = pd.to_numeric(forecast["current_rank"], errors="coerce")
    if sorted(ranks.tolist()) != list(range(1, int(cfg.team_count) + 1)):
        raise ValueError("forecast_summary current_rank must be a complete permutation")

    normalized_games = _normalize_ids(_load_team_games(team_games))
    if set(normalized_games["team_id"]) - set(current["team_id"]):
        raise ValueError("team_games contains a team outside current_standings")
    return frames, normalized_games, manifest, payload


def _clean(value: object) -> object:
    if value is None or value is pd.NA or value is pd.NaT:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            pass
    return value


def _title(sheet, text: str, subtitle: str | None = None) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, sheet.max_column))
    cell = sheet.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    if subtitle:
        cell.comment = None


def _style_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    sheet.freeze_panes = "A3"
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row > 1:
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top")


def _format_header(sheet, row: int = 2) -> None:
    for cell in sheet[row]:
        if cell.value is None:
            continue
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    sheet.row_dimensions[row].height = 32


def _number_format_for_header(header: str) -> str | None:
    lower = header.lower()
    if "%" in header or "probability" in lower or "win pct" in lower:
        return "0.0%"
    if lower in {"date", "game date", "cutoff date", "pbpstats snapshot as of"}:
        return "yyyy-mm-dd"
    if lower in {"wins to date", "losses to date"}:
        return "#,##0"
    if lower == "point diff to date":
        return "+0;-0;0"
    if any(token in lower for token in ("netrtg", "margin", "point diff", "strength", "expected rank", "projected wins")):
        return "0.0"
    if lower in {"gp", "w", "l", "rank", "strength rank", "sos rank", "priority"}:
        return "0"
    if lower.startswith("p") and lower[1:].isdigit():
        return "0.0"
    return None


def _write_table(
    sheet,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    *,
    table_name: str,
    freeze_panes: str = "A3",
) -> None:
    for column, header in enumerate(headers, start=1):
        sheet.cell(2, column, header)
    for row_number, row in enumerate(rows, start=3):
        for column, value in enumerate(row, start=1):
            sheet.cell(row_number, column, _clean(value))

    _title(sheet, title)
    _style_sheet(sheet)
    _format_header(sheet)
    sheet.freeze_panes = freeze_panes
    final_column = get_column_letter(len(headers))
    final_row = max(2, 2 + len(rows))
    ref = f"A2:{final_column}{final_row}"
    sheet.auto_filter.ref = ref
    if rows:
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:2"
    sheet.print_area = f"A1:{final_column}{final_row}"

    for column, header in enumerate(headers, start=1):
        number_format = _number_format_for_header(header)
        if number_format:
            for cell in sheet.iter_cols(
                min_col=column, max_col=column, min_row=3, max_row=final_row
            ):
                for item in cell:
                    item.number_format = number_format
                    item.alignment = Alignment(horizontal="right", vertical="top")
    _set_widths(sheet, min_width=9, max_width=38)


def _set_widths(sheet, *, min_width: int = 8, max_width: int = 42) -> None:
    for cells in sheet.iter_cols(min_row=2):
        max_length = 0
        for cell in cells:
            if cell.value is None:
                continue
            value = str(cell.value)
            max_length = max(max_length, max((len(line) for line in value.splitlines()), default=0))
        sheet.column_dimensions[get_column_letter(cells[0].column)].width = min(
            max_width, max(min_width, max_length + 2)
        )


def _team_context(frames: Mapping[str, pd.DataFrame]) -> pd.DataFrame:
    current = frames["current_standings"].copy()
    strength = frames["team_strength"].copy()
    forecast = frames["forecast_summary"].copy()
    head_to_head = frames["head_to_head"].copy()
    abbreviations = current.set_index("team_id")["team_abbreviation"]

    strength = strength.sort_values(
        ["composite_strength", "team_id"], ascending=[False, True], kind="stable"
    ).reset_index(drop=True)
    strength["strength_rank"] = range(1, len(strength) + 1)
    strength_enrichment_columns = [
        column
        for column in strength.columns
        if column == "team_id" or column not in current.columns
    ]
    context = current.merge(
        strength[strength_enrichment_columns],
        on="team_id",
        how="left",
        validate="one_to_one",
    )
    context = context.merge(
        forecast[
            ["team_id", "playoff_probability", "expected_final_rank"]
        ],
        on="team_id",
        how="left",
        validate="one_to_one",
    )

    h2h_strings: dict[str, str] = {}
    for team_id, rows in head_to_head.groupby("team_id", sort=False):
        parts = []
        for row in rows.sort_values("opponent_id", kind="stable").itertuples():
            opponent = abbreviations.get(row.opponent_id, row.opponent_id)
            parts.append(
                f"vs {opponent}: {int(row.wins)}-{int(row.losses)} "
                f"({float(row.point_differential):+.0f})"
            )
        h2h_strings[team_id] = "; ".join(parts) if parts else "No completed H2H games"

    context["h2h_context"] = context["team_id"].map(h2h_strings)
    context["seed_strength_gap"] = context["current_rank"] - context["strength_rank"]
    return context.sort_values(["current_rank", "team_id"], kind="stable")


def _current_standings_sheet(workbook: Workbook, frames: Mapping[str, pd.DataFrame]) -> None:
    sheet = workbook.create_sheet("Current Standings")
    context = _team_context(frames)
    headers = [
        "Rank",
        "Team",
        "Team ID",
        "GP",
        "W",
        "L",
        "Win %",
        "GB",
        "Home",
        "Road",
        "Last 10",
        "Streak",
        "Conference",
        "Current .500+",
        "Point Diff",
        "Season NetRtg",
        "Recent NetRtg",
        "Composite Strength",
        "Strength Rank",
        "Seed vs Strength Gap",
        "H2H Context",
        "Playoff %",
        "Expected Rank",
    ]
    rows = []
    for row in context.itertuples():
        rows.append(
            [
                row.current_rank,
                row.team_abbreviation,
                str(row.team_id),
                row.games_played,
                row.wins,
                row.losses,
                row.win_pct,
                row.games_back,
                row.home_record,
                row.road_record,
                row.last10_record,
                row.current_streak_label,
                row.conference_record,
                row.record_vs_current_500_plus,
                row.point_differential,
                row.season_net_rating,
                row.recent_net_rating,
                row.composite_strength,
                row.strength_rank,
                row.seed_strength_gap,
                row.h2h_context,
                row.playoff_probability,
                row.expected_final_rank,
            ]
        )
    _write_table(
        sheet,
        "Current Standings & Strength Context",
        headers,
        rows,
        table_name="CurrentStandingsTable",
        freeze_panes="D3",
    )
    header_columns = {header: index for index, header in enumerate(headers, start=1)}
    sheet.column_dimensions[get_column_letter(header_columns["H2H Context"])].width = 34
    sheet.column_dimensions[get_column_letter(header_columns["Current .500+"])].width = 18
    for row in range(3, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 72
        for column in (
            header_columns["H2H Context"],
            header_columns["Current .500+"],
        ):
            sheet.cell(row, column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    for column in (header_columns["Win %"], header_columns["Playoff %"]):
        sheet.conditional_formatting.add(
            f"{get_column_letter(column)}3:{get_column_letter(column)}{sheet.max_row}",
            ColorScaleRule(
                start_type="min", start_color=PALE_RED,
                mid_type="percentile", mid_value=50, mid_color=PALE_GOLD,
                end_type="max", end_color=PALE_GREEN,
            ),
        )


def _remaining_schedule_sheet(workbook: Workbook, frames: Mapping[str, pd.DataFrame]) -> None:
    sheet = workbook.create_sheet("Remaining Schedule")
    current = frames["current_standings"]
    forecast = frames["forecast_summary"]
    matchups = frames["matchup_probabilities"].copy()
    leverage = frames["playoff_leverage_games"]
    abbreviations = current.set_index("team_id")["team_abbreviation"]
    ranks = current.set_index("team_id")["current_rank"]
    playoff = forecast.set_index("team_id")["playoff_probability"]
    display = matchups.merge(
        leverage[
            [
                "game_id",
                "leverage_score",
                "leverage_label",
                "direct_h2h_tiebreak_flag",
            ]
        ],
        on="game_id",
        how="left",
        validate="one_to_one",
    ).sort_values(["game_date", "game_id"], kind="stable")
    headers = [
        "Date",
        "Game ID",
        "Away",
        "Away Rank",
        "Away Playoff %",
        "Home",
        "Home Rank",
        "Home Playoff %",
        "Away Win %",
        "Home Win %",
        "Expected Home Margin",
        "Away Rest",
        "Home Rest",
        "Away B2B",
        "Home B2B",
        "Leverage",
        "Leverage Label",
        "H2H Flag",
        "Checkpoint",
    ]
    rows = []
    for row in display.itertuples():
        label = row.leverage_label if pd.notna(row.leverage_label) else "Unavailable"
        rows.append(
            [
                pd.Timestamp(row.game_date),
                str(row.game_id),
                abbreviations.get(row.away_id, row.away_id),
                ranks.get(row.away_id),
                playoff.get(row.away_id),
                abbreviations.get(row.home_id, row.home_id),
                ranks.get(row.home_id),
                playoff.get(row.home_id),
                row.away_win_probability,
                row.home_win_probability,
                row.expected_home_margin,
                row.away_rest_days,
                row.home_rest_days,
                row.away_b2b,
                row.home_b2b,
                row.leverage_score,
                label,
                row.direct_h2h_tiebreak_flag,
                f"Pregame: {label} leverage; revisit standings at halftime and Q4.",
            ]
        )
    _write_table(
        sheet,
        "Remaining Schedule & Playoff Leverage",
        headers,
        rows,
        table_name="RemainingScheduleTable",
    )
    sheet.column_dimensions["S"].width = 38
    for row in range(3, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 36
        sheet.cell(row, 19).alignment = Alignment(wrap_text=True, vertical="top")
    if rows:
        sheet.conditional_formatting.add(
            f"P3:P{sheet.max_row}",
            ColorScaleRule(
                start_type="min", start_color=PALE_GREEN,
                mid_type="percentile", mid_value=50, mid_color=PALE_GOLD,
                end_type="max", end_color=PALE_RED,
            ),
        )


def _forecast_sheet(workbook: Workbook, frames: Mapping[str, pd.DataFrame], cfg: object) -> None:
    sheet = workbook.create_sheet("Forecast")
    forecast = frames["forecast_summary"].sort_values(
        ["current_rank", "team_id"], kind="stable"
    )
    rank_headers = [f"Rank {rank} %" for rank in range(1, int(cfg.team_count) + 1)]
    headers = [
        "Current Rank",
        "Team",
        "Team ID",
        "Projected Wins",
        "P10",
        "P50",
        "P90",
        "Playoff %",
        "Top-4 %",
        "Home-Court %",
        "Expected Rank",
        *rank_headers,
        "OUT %",
        "SOS Rank",
    ]
    rows = []
    for row in forecast.to_dict("records"):
        rows.append(
            [
                row["current_rank"],
                row["team_abbreviation"],
                str(row["team_id"]),
                row["projected_wins_mean"],
                row["wins_p10"],
                row["wins_p50"],
                row["wins_p90"],
                row["playoff_probability"],
                row["top4_probability"],
                row["home_court_probability"],
                row["expected_final_rank"],
                *[
                    row[f"rank_{rank}_probability"]
                    for rank in range(1, int(cfg.team_count) + 1)
                ],
                row["out_probability"],
                row["remaining_sos_rank"],
            ]
        )
    _write_table(
        sheet,
        "Forecast Distribution — Supplied Model Outputs",
        headers,
        rows,
        table_name="ForecastTable",
        freeze_panes="D3",
    )
    first_probability = 8
    last_probability = 11 + int(cfg.team_count) + 1
    for column in range(first_probability, last_probability + 1):
        letter = get_column_letter(column)
        sheet.conditional_formatting.add(
            f"{letter}3:{letter}{sheet.max_row}",
            ColorScaleRule(
                start_type="min", start_color=WHITE,
                mid_type="percentile", mid_value=50, mid_color=SKY,
                end_type="max", end_color=BLUE,
            ),
        )


def _broadcast_insights_sheet(workbook: Workbook, frames: Mapping[str, pd.DataFrame]) -> None:
    sheet = workbook.create_sheet("Broadcast Insights")
    frame = frames["broadcast_insights"].sort_values(
        ["priority", "category", "team_or_race"], kind="stable"
    )
    columns = [
        ("Priority", "priority"),
        ("Category", "category"),
        ("Team / Race", "team_or_race"),
        ("Data Point", "data_point"),
        ("Quick Read", "quick_read_snippet"),
        ("Pregame Trigger", "pregame_trigger"),
        ("Halftime Checkpoint", "halftime_checkpoint"),
        ("Fourth-Quarter Checkpoint", "fourth_quarter_checkpoint"),
        ("Why It Matters", "why_it_matters"),
        ("Source Fields", "source_fields"),
    ]
    rows = [[row[source] for _, source in columns] for row in frame.to_dict("records")]
    _write_table(
        sheet,
        "Deterministic Broadcast Insights",
        [header for header, _ in columns],
        rows,
        table_name="BroadcastInsightsTable",
    )
    for column in range(5, 11):
        sheet.column_dimensions[get_column_letter(column)].width = 36
        for cell in sheet.iter_cols(
            min_col=column, max_col=column, min_row=3, max_row=sheet.max_row
        ):
            for item in cell:
                item.alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(3, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 48


def _team_games_sheet(workbook: Workbook, team_games: pd.DataFrame) -> None:
    sheet = workbook.create_sheet("Team Games Source")
    preferred = [
        "season",
        "game_id",
        "game_date",
        "team_abbreviation",
        "team_id",
        "opponent_abbreviation",
        "opponent_id",
        "home_away",
        "win",
        "loss",
        "points_for",
        "points_against",
        "margin",
        "wins_to_date",
        "losses_to_date",
        "win_pct_to_date",
        "point_diff_to_date",
        "source_game_completed",
        "source_team_box_path",
    ]
    columns = [column for column in preferred if column in team_games.columns]
    columns.extend(column for column in team_games.columns if column not in columns)
    ordered = team_games.sort_values(
        ["game_date", "game_id", "team_id"], kind="stable"
    )

    def header(column: str) -> str:
        overrides = {
            "game_id": "Game ID",
            "team_id": "Team ID",
            "opponent_id": "Opponent ID",
            "game_date": "Date",
        }
        return overrides.get(column, column.replace("_", " ").title())

    _write_table(
        sheet,
        "Normalized Team-Game Source (Directional Audit Rows)",
        [header(column) for column in columns],
        [[row[column] for column in columns] for row in ordered.to_dict("records")],
        table_name="TeamGamesSourceTable",
        freeze_panes="F3",
    )
    for column in columns:
        if column.endswith("_id") or column == "game_id":
            index = columns.index(column) + 1
            for cell in sheet.iter_cols(
                min_col=index, max_col=index, min_row=3, max_row=sheet.max_row
            ):
                for item in cell:
                    item.number_format = "@"


def _historical_summary(history: pd.DataFrame) -> str:
    available = history.loc[history["availability_status"].eq("available")]
    if available.empty:
        return "Historical context unavailable for this run."
    preferred = available.loc[
        available["metric"].isin(("final_qualifier_wins", "final_8th_place_wins"))
    ]
    row = (preferred if not preferred.empty else available).iloc[0]
    season = "multi-season" if pd.isna(row.get("season")) else str(int(row["season"]))
    value = row.get("value")
    value_text = "n/a" if pd.isna(value) else f"{float(value):.1f}"
    return f"{season}: {row['metric']} = {value_text}."


def _section(sheet, row: int, column: int, label: str, width: int = 4) -> None:
    sheet.merge_cells(
        start_row=row,
        start_column=column,
        end_row=row,
        end_column=column + width - 1,
    )
    cell = sheet.cell(row, column, label)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 22


def _dashboard_sheet(workbook: Workbook, frames: Mapping[str, pd.DataFrame], cfg: object, manifest: dict) -> None:
    sheet = workbook.create_sheet("Dashboard")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 95
    sheet.merge_cells("A1:H1")
    sheet["A1"] = f"WNBA {cfg.season} Standings & Playoff Forecast"
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells("A2:H2")
    sheet["A2"] = (
        f"Cutoff {manifest['cutoff_date']}  |  Model {manifest['model_version']}  |  "
        f"{int(manifest['simulation_count']):,} simulations"
    )
    sheet["A2"].font = SUBTLE_FONT

    forecast = frames["forecast_summary"].sort_values(
        ["current_rank", "team_id"], kind="stable"
    )
    qualifier = int(cfg.playoff_qualifiers)
    _section(sheet, 4, 1, "Current Leader")
    sheet["A5"] = forecast.iloc[0].team_abbreviation
    sheet["A5"].font = Font(name="Aptos Display", size=16, bold=True, color=NAVY)

    cutline_label = f"{cfg.season} Playoff Cutline ({qualifier}th / {qualifier + 1}th)"
    if qualifier == 1:
        cutline_label = f"{cfg.season} Playoff Cutline (1st / 2nd)"
    _section(sheet, 4, 5, cutline_label)
    in_team = forecast.loc[forecast["current_rank"].eq(qualifier)].iloc[0]
    out_team = forecast.loc[forecast["current_rank"].eq(qualifier + 1)].iloc[0]
    sheet["E5"] = (
        f"{in_team.team_abbreviation} {in_team.playoff_probability:.1%}  |  "
        f"{out_team.team_abbreviation} {out_team.playoff_probability:.1%}"
    )

    _section(sheet, 7, 1, "Projected Playoff Field", width=8)
    projected = forecast.sort_values(
        ["expected_final_rank", "team_id"], kind="stable"
    ).head(qualifier)
    sheet.merge_cells("A8:H8")
    sheet["A8"] = "  •  ".join(
        f"{row.team_abbreviation} ({row.playoff_probability:.0%})"
        for row in projected.itertuples()
    )

    _section(sheet, 10, 1, "Top-4 / Home-Court Race", width=4)
    top_limit = min(4, qualifier)
    top = forecast.sort_values(
        ["home_court_probability", "team_id"], ascending=[False, True], kind="stable"
    ).head(top_limit)
    sheet.merge_cells("A11:D11")
    sheet["A11"] = "  •  ".join(
        f"{row.team_abbreviation} {row.home_court_probability:.0%}"
        for row in top.itertuples()
    )

    rank_columns = [f"rank_{rank}_probability" for rank in range(1, int(cfg.team_count) + 1)]
    uncertainty = forecast[rank_columns].max(axis=1).rsub(1.0)
    uncertain = forecast.loc[uncertainty.sort_values(ascending=False, kind="stable").index[0]]
    _section(sheet, 10, 5, "Most Uncertain Seed", width=4)
    sheet.merge_cells("E11:H11")
    sheet["E11"] = (
        f"{uncertain.team_abbreviation}: expected rank {uncertain.expected_final_rank:.2f}; "
        f"largest rank share {uncertain[rank_columns].max():.1%}"
    )

    _section(sheet, 13, 1, "Remaining SOS Extremes", width=4)
    sos = forecast.dropna(subset=["remaining_sos_rank"]).sort_values(
        ["remaining_sos_rank", "team_id"], kind="stable"
    )
    sheet.merge_cells("A14:D14")
    if sos.empty:
        sheet["A14"] = "No remaining schedule."
    else:
        sheet["A14"] = (
            f"Hardest: {sos.iloc[0].team_abbreviation} (rank {int(sos.iloc[0].remaining_sos_rank)})  |  "
            f"Easiest: {sos.iloc[-1].team_abbreviation} (rank {int(sos.iloc[-1].remaining_sos_rank)})"
        )

    _section(sheet, 13, 5, "Historical Cutline Comparison", width=4)
    sheet.merge_cells("E14:H14")
    sheet["E14"] = _historical_summary(frames["historical_context"])

    _section(sheet, 16, 1, "Highest-Leverage Games", width=8)
    leverage = frames["playoff_leverage_games"].sort_values(
        ["leverage_score", "game_date", "game_id"],
        ascending=[False, True, True],
        kind="stable",
    ).head(5)
    current = frames["current_standings"]
    abbreviations = current.set_index("team_id")["team_abbreviation"]
    start = 17
    if leverage.empty:
        sheet.cell(start, 1, "No remaining games.")
    else:
        for offset, row in enumerate(leverage.itertuples()):
            sheet.merge_cells(
                start_row=start + offset,
                start_column=1,
                end_row=start + offset,
                end_column=8,
            )
            sheet.cell(
                start + offset,
                1,
                f"{pd.Timestamp(row.game_date).date().isoformat()}  "
                f"{abbreviations.get(row.away_id, row.away_id)} at "
                f"{abbreviations.get(row.home_id, row.home_id)}  —  "
                f"{row.leverage_score:.0f} ({row.leverage_label})",
            )

    story_row = start + max(2, len(leverage)) + 1
    _section(sheet, story_row, 1, "Headline Storylines", width=8)
    insights = frames["broadcast_insights"].sort_values(
        ["priority", "category"], kind="stable"
    ).head(5)
    for offset, row in enumerate(insights.itertuples(), start=1):
        sheet.merge_cells(
            start_row=story_row + offset,
            start_column=1,
            end_row=story_row + offset,
            end_column=8,
        )
        sheet.cell(story_row + offset, 1, f"{int(row.priority)}. {row.quick_read_snippet}")

    for column in range(1, 9):
        sheet.column_dimensions[get_column_letter(column)].width = 16
    for row in range(3, sheet.max_row + 1):
        for cell in sheet[row]:
            if cell.value is not None and cell.fill.fill_type is None:
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A4"
    sheet.print_area = f"A1:H{sheet.max_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1


def _model_notes_sheet(workbook: Workbook, frames: Mapping[str, pd.DataFrame], cfg: object, manifest: dict) -> None:
    sheet = workbook.create_sheet("Model Notes")
    history_available = frames["historical_context"]["availability_status"].eq("available").any()
    rows = [
        ["Cutoff date", manifest.get("cutoff_date")],
        ["Season", manifest.get("season")],
        ["Model version", manifest.get("model_version")],
        ["Simulation count", manifest.get("simulation_count")],
        ["Conditional simulation count", manifest.get("conditional_simulation_count")],
        ["Random seed", manifest.get("random_seed")],
        ["Git SHA", manifest.get("git_sha") or "Unavailable"],
        ["Official tiebreak fallback count", manifest.get("official_tiebreak_fallback_count")],
        ["Tiebreak fallback", "Stable normalized team ID is used only after every official criterion remains tied; it is not an official WNBA tiebreak."],
        ["PBPStats status", manifest.get("pbpstats_enrichment_status")],
        ["History status", "available" if history_available else "unavailable"],
        ["History seasons used", ", ".join(map(str, manifest.get("history_seasons_used", []))) or "None"],
        ["Qualifier definition", f"Top {cfg.playoff_qualifiers} of {cfg.team_count} league-wide teams qualify."],
        ["Home-court definition", "Top four probability is also reported as home-court probability for this V1 forecast."],
        ["SOS definition", "Remaining SOS rank is ordered by mean supplied opponent playoff probability; rank 1 is hardest."],
        ["Strength definition", "Composite strength is explanatory; predictive NetRtg drives matchup expected margin upstream."],
        ["Current standings source", CURRENT_STANDINGS_METHOD],
        ["Current vs simulated-final .500+", CURRENT_500_METHOD],
        ["Probability source", "All probabilities come from the validated machine-readable forecast bundle; this workbook does not resimulate or recalculate them."],
        ["Linked standings leader", "='Current Standings'!B3"],
        ["Caveat", "Forecasts are probabilistic, source-snapshot dependent, and subject to official standings/tiebreak updates."],
        ["Season config SHA-256", manifest.get("season_config_sha256")],
        ["Model config SHA-256", manifest.get("model_config_sha256")],
    ]
    rows.extend(
        [
            f"Source: {source.get('name', 'unknown')}",
            f"{source.get('path', 'unavailable')} | sha256 "
            f"{source.get('sha256', 'unavailable')} | "
            f"{source.get('size_bytes', 'unknown')} bytes",
        ]
        for source in manifest.get("source_files", [])
    )
    _write_table(
        sheet,
        "Model Notes, Definitions & Provenance",
        ["Item", "Value"],
        rows,
        table_name="ModelNotesTable",
    )
    sheet.auto_filter.ref = None
    for cell in sheet[3]:
        cell.fill = PatternFill("solid", fgColor=PALE_GOLD)
        cell.font = Font(name="Aptos", size=10, bold=True, color=NAVY)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 72
    sheet.page_setup.fitToHeight = 1
    for row in range(3, sheet.max_row + 1):
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        if len(str(sheet.cell(row, 2).value or "")) > 120:
            sheet.row_dimensions[row].height = 42


def _build_workbook(
    frames: Mapping[str, pd.DataFrame],
    team_games: pd.DataFrame,
    cfg: object,
    manifest: dict,
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _dashboard_sheet(workbook, frames, cfg, manifest)
    _current_standings_sheet(workbook, frames)
    _remaining_schedule_sheet(workbook, frames)
    _forecast_sheet(workbook, frames, cfg)
    _broadcast_insights_sheet(workbook, frames)
    _team_games_sheet(workbook, team_games)
    _model_notes_sheet(workbook, frames, cfg, manifest)
    if tuple(workbook.sheetnames) != SHEET_ORDER:
        raise RuntimeError("workbook sheet order drifted from the public contract")
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    return workbook


def _atomic_save(workbook: Workbook, final_path: Path) -> None:
    final_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{final_path.name}.", suffix=".tmp", dir=final_path.parent
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        with temporary_path.open("rb") as handle:
            os.fsync(handle.fileno())
        os.replace(temporary_path, final_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def render_excel(
    processed_root: str | Path,
    team_games: pd.DataFrame | str | Path,
    *,
    cfg: object,
) -> Path:
    """Render one validated bundle without calculating forecast probabilities."""

    root = Path(processed_root).expanduser().resolve()
    frames, normalized_games, manifest, _payload = _load_inputs(root, team_games, cfg)
    workbook = _build_workbook(frames, normalized_games, cfg, manifest)

    output_root = Path(cfg.output_root).expanduser()
    if not output_root.is_absolute():
        raise ValueError("configured output_root must be absolute before rendering")
    final_path = (
        output_root
        / "deliverables"
        / f"season={cfg.season}"
        / "latest"
        / WORKBOOK_FILENAME
    )
    _atomic_save(workbook, final_path)
    return final_path


__all__ = ["SHEET_ORDER", "WORKBOOK_FILENAME", "render_excel"]
